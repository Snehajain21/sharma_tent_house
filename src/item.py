import os

from datetime import datetime

from openpyxl import Workbook

from helpers import (
    read_positive_int
)

from storage import (
    save_items
)


# GET CURRENT TIME
def get_current_time():

    return datetime.now().strftime(
        "%d-%m-%Y %H:%M:%S"
    )


# VIEW ITEM DETAILS
def view_item_details(item):

    print("\n========== ITEM DETAILS ==========")

    print("Item ID:",
          item["item_id"])

    print("Item Name:",
          item["item_name"])

    print("Total Quantity:",
          item["total_quantity"])

    print("Rental Price:",
          item["rental_price"])

    print("Last Updated:",
          item["last_updated"])


# ADD NEW ITEM
def add_new_item(items):

    item_name = input(
        "Enter new item name: "
    ).strip()

    # NORMALIZE NEW ITEM NAME
    normalized_new_name = " ".join(
        item_name.lower().split()
    )

    # CHECK IF ITEM ALREADY EXISTS
    for item in items:

        normalized_existing_name = " ".join(
            item["item_name"].lower().split()
        )

        if (
            normalized_existing_name
            == normalized_new_name
        ):

            print(
                f"{item_name} already exists."
            )

            return

    total_quantity = read_positive_int(
        "Enter item quantity: "
    )

    rental_price = read_positive_int(
        "Enter rental price: "
    )

    # FIND NEXT UNIQUE ITEM ID
    max_id = 100

    for item in items:

        current_id = int(
            item["item_id"][1:]
        )

        if current_id > max_id:

            max_id = current_id

    item_id = "I" + str(max_id + 1)

    new_item = {

        "item_id": item_id,

        "item_name": item_name,

        "total_quantity": total_quantity,

        "rental_price": rental_price,

        "last_updated": "Not updated yet"
    }

    items.append(new_item)

    save_items(items)

    print(
        f"{item_name} added successfully with quantity {total_quantity}."
    )


# DELETE ITEM
def delete_item(item, items):

    confirmation = input(
        f"Are you sure you want to delete '{item['item_name']}'? (yes/no): "
    ).strip().lower()

    if confirmation != "yes":

        print("Delete operation cancelled.")
        return

    items.remove(item)

    save_items(items)

    print(
        f"{item['item_name']} deleted successfully."
    )


# ADD ITEM QUANTITY
def add_item_quantity(item, items):

    quantity = read_positive_int(
        "Enter quantity to add: "
    )

    item["total_quantity"] += quantity

    item["last_updated"] = get_current_time()

    save_items(items)

    print(
        f"{item['item_name']} quantity updated to {item['total_quantity']}."
    )


# REMOVE ITEM QUANTITY
def remove_item_quantity(item, items):

    quantity = read_positive_int(
        "Enter quantity to remove: "
    )

    if quantity > item["total_quantity"]:

        print("Not enough quantity available.")
        return

    item["total_quantity"] -= quantity

    item["last_updated"] = get_current_time()

    save_items(items)

    print(
        f"{item['item_name']} quantity updated to {item['total_quantity']}."
    )


# UPDATE ITEM NAME
# UPDATE ITEM NAME
def update_item_name(item, items):

    new_name = input(
        "Enter new item name: "
    ).strip()

    if new_name == "":

        print("Item name cannot be empty.")
        return

    # NORMALIZE NEW NAME
    normalized_new_name = " ".join(
        new_name.lower().split()
    )

    # CHECK DUPLICATE NAMES
    for existing_item in items:

        normalized_existing_name = " ".join(
            existing_item["item_name"]
            .lower()
            .split()
        )

        if (
            existing_item != item
            and normalized_existing_name
            == normalized_new_name
        ):

            print(
                f"{new_name} already exists."
            )

            return

    item["item_name"] = new_name

    item["last_updated"] = get_current_time()

    save_items(items)

    print(
        f"Item name updated to {item['item_name']}."
    )


# UPDATE RENTAL PRICE
def update_rental_price(item, items):

    new_price = read_positive_int(
        "Enter new rental price: "
    )

    item["rental_price"] = new_price

    item["last_updated"] = get_current_time()

    save_items(items)

    print(
        f"{item['item_name']} rental price updated to {item['rental_price']}."
    )
# EXPORT ALL ITEMS TO EXCEL
def export_items_to_excel(items):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Inventory"

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    # HEADER STYLE
    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=12
    )

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin_border = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin")
    )

    # HEADERS
    headers = [

        "Item ID",

        "Item Name",

        "Total Quantity",

        "Rental Price",

        "Last Updated"
    ]

    sheet.append(headers)

    # STYLE HEADER ROW
    for cell in sheet[1]:

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = center_alignment

        cell.border = thin_border

    # ITEM DATA
    for item in items:

        sheet.append([

            item["item_id"],

            item["item_name"],

            item["total_quantity"],

            item["rental_price"],

            item["last_updated"]
        ])

    # STYLE DATA ROWS
    for row in sheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = center_alignment

            cell.border = thin_border

    # AUTO COLUMN WIDTH
    for column in sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:

                pass

        adjusted_width = max_length + 5

        sheet.column_dimensions[
            column_letter
        ].width = adjusted_width

    # SINGLE FILE NAME
    file_name = "inventory.xlsx"

    # SAVE AND OPEN FILE
    try:

        workbook.save(file_name)

        print(
            "Opening inventory file..."
        )

        os.startfile(file_name)

    except PermissionError:

        print(
            "Please close the opened inventory Excel file and try again."
        )