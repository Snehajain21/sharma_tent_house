import os

from openpyxl import Workbook

from datetime import datetime

from helpers import (
    search_item,
    read_positive_int
)

from storage import (
    load_bookings,
    save_bookings
)

# GENERATE BOOKING ID
def generate_booking_id():

    bookings = load_bookings()

    max_id = 100

    for booking in bookings:

        current_id = int(
            booking["booking_id"][1:]
        )

        if current_id > max_id:

            max_id = current_id

    return "B" + str(max_id + 1)


# READ VALID DATE
def read_valid_date(date_type):

    while True:

        print(
            f"\nEnter {date_type}"
        )

        day = read_positive_int(
            "Enter day: "
        )

        month = read_positive_int(
            "Enter month: "
        )

        year = read_positive_int(
            "Enter year: "
        )

        date_text = (
            f"{year}-{month:02d}-{day:02d}"
        )

        try:

            datetime.strptime(
                date_text,
                "%Y-%m-%d"
            )

            return date_text

        except ValueError:

            print(
                f"\n❌ Invalid date: {date_text}"
            )

            print(
                "This date does not exist in the calendar."
            )


# CHECK DATE OVERLAP
def dates_overlap(

    booking_dispatch_date,

    booking_return_date,

    request_dispatch_date,

    request_return_date

):

    booking_dispatch_date = datetime.strptime(
        booking_dispatch_date,
        "%Y-%m-%d"
    )

    booking_return_date = datetime.strptime(
        booking_return_date,
        "%Y-%m-%d"
    )

    request_dispatch_date = datetime.strptime(
        request_dispatch_date,
        "%Y-%m-%d"
    )

    request_return_date = datetime.strptime(
        request_return_date,
        "%Y-%m-%d"
    )

    return not (

        request_return_date
        < booking_dispatch_date

        or

        request_dispatch_date
        > booking_return_date
    )


# CHECK AVAILABILITY
def check_availability(

    item,

    dispatch_date,

    expected_return_date

):

    bookings = load_bookings()

    reserved_quantity = 0

    for booking in bookings:

        if booking["item_id"] != item["item_id"]:

            continue

        overlap = dates_overlap(

            booking["dispatch_date"],

            booking["expected_return_date"],

            dispatch_date,

            expected_return_date
        )

        if overlap:

            reserved_quantity += (
                booking["quantity"]
            )

    available_quantity = (

        item["total_quantity"]

        - reserved_quantity
    )

    return available_quantity

# CHECK AVAILABILITY FOR UPDATE
def check_availability_for_update(

    booking_to_update,

    item,

    dispatch_date,

    expected_return_date

):

    bookings = load_bookings()

    reserved_quantity = 0

    for booking in bookings:

        # SKIP CURRENT BOOKING
        if (

            booking["booking_id"]

            ==

            booking_to_update["booking_id"]

        ):

            continue

        # SKIP DIFFERENT ITEMS
        if (

            booking["item_id"]

            !=

            item["item_id"]

        ):

            continue

        overlap = dates_overlap(

            booking["dispatch_date"],

            booking["expected_return_date"],

            dispatch_date,

            expected_return_date
        )

        if overlap:

            reserved_quantity += (
                booking["quantity"]
            )

    available_quantity = (

        item["total_quantity"]

        - reserved_quantity
    )

    return available_quantity


# VALIDATE BOOKING DATES
def validate_booking_dates(

    dispatch_date,

    expected_return_date

):

    dispatch_datetime = datetime.strptime(

        dispatch_date,

        "%Y-%m-%d"
    )

    return_datetime = datetime.strptime(

        expected_return_date,

        "%Y-%m-%d"
    )

    today = datetime.today()

    if dispatch_datetime.date() < today.date():

        print(
            f"\n❌ Invalid dispatch date: {dispatch_date}"
        )

        print(
            "Past dates cannot be used for new bookings."
        )

        return False

    if return_datetime < dispatch_datetime:

        print(
            "\n❌ Invalid booking period."
        )

        print(
            f"Dispatch Date: {dispatch_date}"
        )

        print(
            f"Expected Return Date: {expected_return_date}"
        )

        print(
            "The expected return date cannot be earlier than the dispatch date."
        )

        return False

    return True

# CREATE BOOKING
def create_booking(items):

    item = search_item(items)

    if item is None:

        return

    quantity = read_positive_int(
        "Enter quantity: "
    )

    dispatch_date = read_valid_date(
        "Dispatch Date"
    )

    expected_return_date = read_valid_date(
        "Expected Return Date"
    )

    if not validate_booking_dates(

        dispatch_date,

        expected_return_date

    ):

        return

    # AVAILABILITY CHECK
    available_quantity = check_availability(

        item,

        dispatch_date,

        expected_return_date
    )

    if quantity > available_quantity:

        print(
            "\n❌ Booking Rejected."
        )

        print(
            f"Available Quantity: {available_quantity}"
        )

        print(
            f"Requested Quantity: {quantity}"
        )

        return

    booking = {

        "booking_id": generate_booking_id(),

        "item_id": item["item_id"],

        "quantity": quantity,

        "dispatch_date": dispatch_date,

        "expected_return_date": expected_return_date
    }

    bookings = load_bookings()

    bookings.append(
        booking
    )

    save_bookings(
        bookings
    )

    print(
        f"\n✅ Booking {booking['booking_id']} created successfully."
    )

    print(
        "\n========== BOOKING SUMMARY =========="
    )

    print(
        f"Booking ID: {booking['booking_id']}"
    )

    print(
        f"Item ID: {booking['item_id']}"
    )

    print(
        f"Item Name: {item['item_name']}"
    )

    print(
        f"Quantity: {booking['quantity']}"
    )

    print(
        f"Dispatch Date: {booking['dispatch_date']}"
    )

    print(
        f"Expected Return Date: {booking['expected_return_date']}"
    )

# VIEW ALL BOOKINGS
def view_all_bookings():

    bookings = load_bookings()

    if len(bookings) == 0:

        print(
            "\nNo bookings found."
        )

        return

    export_bookings_to_excel()

# EXPORT BOOKINGS TO EXCEL
def export_bookings_to_excel():

    bookings = load_bookings()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Bookings"

    from storage import load_items

    items = load_items()

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    # HEADER STYLE
    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=12
    )

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin_border = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin")
    )

    # HEADERS
    headers = [

        "Booking ID",

        "Item ID",

        "Item Name",

        "Quantity",

        "Dispatch Date",

        "Expected Return Date"
    ]

    sheet.append(
        headers
    )

    # STYLE HEADER
    for cell in sheet[1]:

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = center_alignment

        cell.border = thin_border

    # DATA ROWS
    for booking in bookings:

        item_name = "Unknown Item"

        for item in items:

            if item["item_id"] == booking["item_id"]:

                item_name = item["item_name"]

                break

        sheet.append([

            booking["booking_id"],

            booking["item_id"],

            item_name,

            booking["quantity"],

            booking["dispatch_date"],

            booking["expected_return_date"]
        ])

    # STYLE DATA
    for row in sheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = center_alignment

            cell.border = thin_border

    # AUTO WIDTH
    for column in sheet.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            try:

                if (
                    len(str(cell.value))
                    > max_length
                ):

                    max_length = len(
                        str(cell.value)
                    )

            except:

                pass

        sheet.column_dimensions[
            column_letter
        ].width = max_length + 5

    file_name = "bookings.xlsx"

    try:

        workbook.save(
            file_name
        )

        print(
            "Opening bookings file..."
        )

        os.startfile(
            file_name
        )

    except PermissionError:

        print(
            "Please close the opened bookings Excel file and try again."
        )

# SEARCH BOOKING
def search_booking(bookings):

    if len(bookings) == 0:

        print(
            "\nNo bookings found."
        )

        return None

    booking_id = input(
        "\nEnter Booking ID: "
    ).strip().upper()

    for booking in bookings:

        if booking["booking_id"] == booking_id:

            return booking

    print(
        "\nBooking not found."
    )

    return None

# VIEW BOOKING DETAILS
def view_booking_details(booking):

    print(
        "\n========== BOOKING DETAILS =========="
    )

    print(
        f"Booking ID: {booking['booking_id']}"
    )

    

    from storage import load_items

    items = load_items()

    item_name = "Unknown Item"

    for item in items:

        if item["item_id"] == booking["item_id"]:

            item_name = item["item_name"]

            break

    print(
        f"Item ID: {booking['item_id']}"
    )

    print(
        f"Item Name: {item_name}"
    )

    

    print(
        f"Quantity: {booking['quantity']}"
    )

    print(
        f"Dispatch Date: {booking['dispatch_date']}"
    )

    print(
        f"Expected Return Date: {booking['expected_return_date']}"
    )



# DELETE BOOKING
def delete_booking():

    bookings = load_bookings()

    booking = search_booking(bookings)

    if booking is None:

        return
    view_booking_details(
    booking
    )

    confirm = input(
        f"\nDelete booking {booking['booking_id']}? (y/n): "
    ).strip().lower()

    if confirm != "y":

        print(
            "\nDeletion cancelled."
        )

        return

    bookings.remove(
        booking
    )

    save_bookings(
        bookings
    )

    print(
        f"\nBooking {booking['booking_id']} deleted successfully."
    )
# UPDATE BOOKING
def update_booking():

    bookings = load_bookings()

    booking = search_booking(
        bookings
    )

    if booking is None:

        return

    view_booking_details(
        booking
    )

    while True:

        print(
            f"\n========== Updating {booking['booking_id']} =========="
        )

        print("1. Update Quantity")

        print("2. Update Dispatch Date")

        print("3. Update Return Date")

        print("4. Exit")

        choice = input(
            "\nEnter choice: "
        )

        if choice == "1":

            old_quantity = booking["quantity"]

            new_quantity = read_positive_int(
                "Enter new quantity: "
            )

            from storage import load_items

            items = load_items()

            item = None

            for current_item in items:

                if (
                    current_item["item_id"]
                    ==
                    booking["item_id"]
                ):

                    item = current_item

                    break

            available_quantity = (

                check_availability_for_update(

                    booking,

                    item,

                    booking["dispatch_date"],

                    booking["expected_return_date"]
                )
            )

            if new_quantity > available_quantity:

                print(
                    "\n❌ Booking update rejected."
                )

                print(
                    f"Available Quantity: {available_quantity}"
                )

                print(
                    f"Requested Quantity: {new_quantity}"
                )

                continue

            confirm = input(

                f"\nUpdate quantity from {old_quantity} to {new_quantity}? (y/n): "

            ).strip().lower()

            if confirm != "y":

                print(
                    "\nUpdate cancelled."
                )

                continue

            booking["quantity"] = (
                new_quantity
            )

            save_bookings(
                bookings
            )

            print(
                f"\nBooking {booking['booking_id']} updated successfully."
            )

            print(
                f"Old Quantity: {old_quantity}"
            )

            print(
                f"New Quantity: {new_quantity}"
            )

        elif choice == "2":

            old_dispatch_date = (
                booking["dispatch_date"]
            )

            new_dispatch_date = read_valid_date(
                "Dispatch Date"
            )

            if not validate_booking_dates(

                new_dispatch_date,

                booking["expected_return_date"]

            ):

                continue

            confirm = input(

                f"\nUpdate dispatch date from {old_dispatch_date} to {new_dispatch_date}? (y/n): "

            ).strip().lower()

            if confirm != "y":

                print(
                    "\nUpdate cancelled."
                )

                continue

            booking["dispatch_date"] = (
                new_dispatch_date
            )

            save_bookings(
                bookings
            )

            print(
                f"\nBooking {booking['booking_id']} updated successfully."
            )

            print(
                f"Old Dispatch Date: {old_dispatch_date}"
            )

            print(
                f"New Dispatch Date: {new_dispatch_date}"
            )

        elif choice == "3":

            old_return_date = (
                booking["expected_return_date"]
            )

            new_return_date = read_valid_date(
                "Expected Return Date"
            )

            if not validate_booking_dates(

                booking["dispatch_date"],

                new_return_date

            ):

                continue

            confirm = input(

                f"\nUpdate return date from {old_return_date} to {new_return_date}? (y/n): "

            ).strip().lower()

            if confirm != "y":

                print(
                    "\nUpdate cancelled."
                )

                continue

            booking["expected_return_date"] = (
                new_return_date
            )

            save_bookings(
                bookings
            )

            print(
                f"\nBooking {booking['booking_id']} updated successfully."
            )

            print(
                f"Old Return Date: {old_return_date}"
            )

            print(
                f"New Return Date: {new_return_date}"
            )

        elif choice == "4":

            print(
                "\nReturning to main menu..."
            )

            break

        else:

            print(
                "\nInvalid choice."
            )